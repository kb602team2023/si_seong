from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
from openpyxl import Workbook,load_workbook
import os.path
from datetime import timedelta,datetime

today = datetime.today()
mymonth = '0'+str(today.month) if len(str(today.month)) ==1 else today.month
yesterday = f"{mymonth}{today.day-1}"
# print(yesterday)

if os.path.isfile('weather.xlsx'):
    wb=load_workbook('weather.xlsx')
else:
    wb=Workbook()

ws = wb.active

# if ws.max_row==1:
#     ws.append(['날짜','도시','날씨','온도'])
ws.append(['날짜','도시','날씨','온도'])
url="https://www.weather.go.kr/w/index.do"

brow=webdriver.Chrome()
brow.get(url)
# print(url)
test = brow.find_element(By.CLASS_NAME,'lay-460').click()
print(test)
# brow.find_element(By.CLASS_NAME,'d1-btn').click()
# print(brow)
# brow.click()
# brow.find_element(By.ClASS_NAME,'on')
# brow.click()

time.sleep(3)
# html=brow.page_source
# print(html)

# html=BeautifulSoup(html,"html.parser")

# a=html.find('dl',{'class','po2_108'})
# print(a)
# city=a.select("dt")
# print(city)
# weather=a.select_one("i")
# tem=a.select("span")
# data=time.strftime("%m" "%d",time.localtime(time.time()))
# result=([data,city.text,weather.text,tem.text])
# print([data,city.text,weather.text,tem.text]) 


