from bs4 import BeautifulSoup
from selenium import webdriver
import time
from openpyxl import Workbook,load_workbook
import os.path
from flask import Flask,render_template

# app=Flask(__name__)

# @app.route("/")
# def weather():
#     return render_template("weather.html")

if os.path.isfile('weather.xlsx'):
    wb=load_workbook('weather.xlsx')
else:
    wb=Workbook()

ws = wb.active

# if ws.max_row==1:
#     ws.append(['날짜','도시','날씨','온도'])
ws.append(['날짜','도시','날씨','온도'])
url="https://www.weather.go.kr/w/index.do"

brow=webdriver.Chrome()
brow.get(url)

time.sleep(3)
html=brow.page_source

html=BeautifulSoup(html,"html.parser")
list=['po_102','po_112','po_108','po_99','po_101','po_90',
'po_104','po_115','po_119','po_131','po_136','po_133',
'po_177','po_146','po_143','po_152','po_138','po_159',
'po_155','po_156','po_165','po_168','po_169','po_184','2714051000']
for i in list:
    a=html.find('dl',{'class',f"{i}"})
    city=a.select_one("dt")
    weather=a.select_one("i")
    tem=a.select_one("span")
    data=time.strftime("%m" "%d",time.localtime(time.time()))
    ws.append([data,city.text,weather.text,tem.text])
    print([data,city.text,weather.text,tem.text])



# wb.save('weather.xlsx')
# wb.close()