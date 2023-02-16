from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
from openpyxl import Workbook,load_workbook
import os.path
from datetime import timedelta,datetime

today = datetime.today()
mymonth = '0'+str(today.month) if len(str(today.month)) ==1 else today.month
yesterday = f"{mymonth}{today.day-1}"
# print(yesterday)

if os.path.isfile('weather.xlsx'):
    wb=load_workbook('weather.xlsx')
else:
    wb=Workbook()

ws = wb.active

# if ws.max_row==1:
#     ws.append(['날짜','도시','날씨','온도'])
ws.append(['날짜','도시','날씨','온도'])
url="https://www.weather.go.kr/w/index.do"

brow=webdriver.Chrome()
brow.get(url)
# print(url)
test = brow.find_element(By.XPATH,'//*[@id="content-body"]/div[3]/h2/a').click()
# print(test)
test = brow.find_element(By.XPATH,'//*[@id="content-body"]/div[3]/div/div/div[1]/ul/li[1]/a/span').click()
# print(test)
# brow.find_element(By.CLASS_NAME,'d1-btn').click()
# print(brow)
# brow.click()
# brow.find_element(By.ClASS_NAME,'on')
# brow.click()

time.sleep(5)
html=brow.page_source
# print(html)

html=BeautifulSoup(html,"html.parser")

a=html.find('dl',{'class','po2_108'})
# print(a)
city=a.select_one("dt")
# print(city.text)
# weather=a.select_one("i")
tem=a.select('span')
# print(tem[0].text)
# print(tem[1])
b=([city.text,tem[0].text,tem[1].text])
print(b)

# tem1=a.select_one("span[1]")
# print(tem1)
# data=time.strftime("%m" "%d",time.localtime(time.time()))
# result=([data,city.text,weather.text,tem.text])
# print([data,city.text,weather.text,tem.text]) 
