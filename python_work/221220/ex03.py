from openpyxl import workbook, load_workbook
import random
class myworkbook:

    def dosaveA1A2(self,a1,a2):
        pass

    # 이 함수 호출 시 sample.xlsx 파일 로딩시 1행 1열 10행 10열 채움
    def writer10c10(self):
        lw=load_workbook("sample.xlsx")
        ws=lw.active
        for x in range(1,11):
            for y in range(1,11):
                ws.cell(row=x,column=y).value = random.randint(0,100)
                print(ws.cell(row=x,column=y))
                print(ws.cell(row=x,column=y).value)
        lw.save("sample.xlsx")
        lw.close()

