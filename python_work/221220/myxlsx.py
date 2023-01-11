import sys
from PyQt5.QtWidgets import *
from PyQt5 import uic
from ex03 import myworkbook,load_workbook

from openpyxl import Workbook

form_class=uic.loadUiType("myxui.ui")[0]

class WindowClass(QMainWindow,form_class):
    def __init__(self):
        super().__init__()
        self.setupUi(self)

        self.e1.setText("1234")
        self.p1.clicked.connect(self.doA)
        self.p2.clicked.connect(self.doB)
        self.p3.clicked.connect(self.do1010)

    def doA(self):
        wb= Workbook()
        wb.save("sample.xlsx")
        wb.close()

    def doB(self):
        lw=load_workbook("sample.xlsx")
        ws=lw.active
        # print(ws["A1"].value)
        # print(ws["A2"].value)
        self.e1.setText(str(ws["A1"].value))
        self.e2.setText(str(ws["A2"].value))
        
        lw.save("sample.xlsx")
        lw.close()

    def do1010(self):
        mub=myworkbook()
        mub.writer10c10()
        
if __name__=="__main__":
    app=QApplication(sys.argv)
    myWindow = WindowClass()
    myWindow.show()
    app.exec_()
