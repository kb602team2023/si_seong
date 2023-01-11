from openpyxl import Workbook
from random import randint

wb=Workbook()
ws=wb.active

ws.append(['번호','국어','영어','수학',])

for i in range(1,11):
    ws.append([i,randint(0,100),randint(0,100),randint(0,100)])

print(ws.max_column)
print(ws.max_row)

a=2
for row in ws.iter_rows(min_row=2):
    total=0
    for inx,cell in enumerate(row):
        if inx !=0:
            total+=int(cell.value)
    ws[f'E{a}']=total
    a+=1
    print()

ws['E1']='총점'


wb.save('test.xlsx')
wb.close()
