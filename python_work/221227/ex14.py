from openpyxl import load_workbook

wb=load_workbook("test.xlsx")
ws=wb.active

ws['F1']='평균'

for row in ws.iter_rows(min_row=2):
    print(row[0].value,row[1].value,row[2].value,row[3].value)
    row[5].value=int(row[4].value/3)


wb.save("test.xlsx")
wb.close()
