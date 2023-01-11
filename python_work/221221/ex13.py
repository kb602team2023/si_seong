from openpyxl import Workbook,load_workbook
class star:
    def makestar(self):
        wb=Workbook()
        ws=wb.active
        for x in range(1,6):
            for y in range(1,x+1):
                ws.cell(row=x,column=y).value = '*'
                
               
               

        wb.save("star.xlsx")
        wb.close()

    def loadstar(self):
        lb = load_workbook("star.xlsx")
        ws=lb.active
        for x in range(1,6):
            for y in range(1,x+1):
                print(ws.cell(row=x,column=y).value,end=" ")

        print()

        lb.close()