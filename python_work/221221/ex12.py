from openpyxl import Workbook,load_workbook
'''
star xlsx 파일 저장하기
star xlsx 파일 불러오기
클래스 명 star로 저장
1매서드 makestra()
2매서드 loadstar()
*
**
***
****
*****
'''
wb=Workbook()
ws=wb.active
for x in range(1,6):
    for y in range(1,x+1):
        ws.cell(row=x,column=y).value = '*'
        
        

wb.save("star.xlsx")
wb.close()

lb = load_workbook("star.xlsx")
ws=lb.active
for x in range(1,6):
    for y in range(1,x+1):
        print(ws.cell(row=x,column=y).value,end=" ")
    print()

lb.close()