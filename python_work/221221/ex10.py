from openpyxl import Workbook,load_workbook
import random

class myworkbook:

    def dosave(self):
        wb = Workbook()
        ws = wb.active
        for x in range(1,10):
            for y in range(1,10):
             ws.cell(row=x,column=y).value = random.randint(0,100)


        wb.save("random.xlsx")
        wb.close()

    def doload(self):
        lb = load_workbook("random.xlsx")
        ws=lb.active
        for x in range(1,10):
            for y in range(1,10):
                print(ws.cell(row=x,column=y).value,end=" ")
            print()

        lb.close()