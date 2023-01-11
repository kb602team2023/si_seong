from myworkbook import myworkbook1

mwb=myworkbook1()

cnt=input("몇행으로 출력할래?")
cnt=int(cnt)+1


mwb.setcnt(cnt)
mwb.dosave()
mwb.doload()

'''
별 내용 관련 작성 자료
from openpyxl import Workbook,load_workbook

# 엑셀에 저장하기
wb=Workbook()
ws=wb.active

cnt=input("몇행으로 출력할래?")
cnt=int(cnt)

for r in range(1,cnt+1):
    for c in range(1,r+1):
        ws.cell(row=r,column=c,value="*"*c)


wb.save("star1.xlsx")
wb.close()

# 엑셀 불러오기
lb=load_workbook("star1.xlsx")
ws=lb.active

for r in range(1,cnt+1):
    for c in range(1,r+1):
        print(ws.cell(row=r,column=c).value,end=" ")
    print()
    
lb.close()
'''