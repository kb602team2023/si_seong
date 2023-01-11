from openpyxl import Workbook,load_workbook

class myworkbook1:
    # 생성자 myworkbook
    def __init__(self):
        self.cnt=1
        print("생성자 호출")

    def setcnt(self,cnt):
        self.cnt=cnt
        print("cnt 변경")

    def dosave(self):
        wb=Workbook()
        ws=wb.active

   
        for r in range(1,self.cnt):
            for c in range(1,r+1):
                ws.cell(row=r,column=c,value="*"*c)


        wb.save("star1.xlsx")
        wb.close()
        print("저장 완료")

    def doload(self):
        lb=load_workbook("star1.xlsx")
        ws=lb.active

        for r in range(1,self.cnt):
            for c in range(1,r+1):
                print(ws.cell(row=r,column=c).value,end=" ")
            print()
            
        lb.close()
        print("불러오기 완료")
