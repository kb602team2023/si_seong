from bs4 import BeautifulSoup
import requests
import re
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

# print(f"bool = { bool(ws.max_row) }")

if ws.max_row ==1 :
    ws.append(['종목명','코드명','날짜','현재가','시가','저가','고가'])

codes = ['033320','192250','335890','012450']

for mycode in codes:
    myhome = f"https://finance.naver.com/item/sise.naver?code={mycode}"
    req = requests.get(myhome)
    req.encoding = "utf-8"

    html = BeautifulSoup(req.content,"html.parser")
    div = html.find('div',{'class','wrap_company'})
    text = re.sub("\W"," ",div.text.strip())
    name = text.split()[0]
    code = text.split()[1]
    date = text.split()[3]+'년'+text.split()[4]+'월'+text.split()[5]+'일'
    print(f"name = {name}")
    print(f"code = {code}")
    print(f"date = {date}")

    nowval = html.find('strong',{id,'_nowVal'})
    c = html.find_all('span',{'class','tah p11'})
    print(f"현재가 {nowval.text}")
    print(f"시가 {c[0].text}")
    print(f"저가 {c[1].text}")
    print(f"고가 {c[2].text}")


    ws.append([name,code,date,nowval.text,c[0].text,c[1].text,c[2].text])

wb.save('jusik.xlsx')
wb.close()