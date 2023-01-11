from openpyxl import Workbook

wb = Workbook()

m1 = wb.create_sheet()
print(wb.sheetnames)

m1.title="my 시트"
m1.sheet_properties.tabColor='ff66ff'

ws1 = wb.create_sheet()
ws2 = wb.create_sheet()
ws3 = wb.create_sheet()
print(wb.sheetnames)

for i in range(1,10):
    wb['Sheet1'][f"A{i}"] = "test"
    wb['Sheet1'][f"B{i}"] = "test"

target = wb.copy_worksheet(wb["Sheet1"])
target.title("copy")

wb.save("sample.xlsx")
wb.close()