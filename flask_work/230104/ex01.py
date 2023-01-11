from flask import Flask,render_template,request
from openpyxl import load_workbook,Workbook
import os.path
import dusrma as ds

app=Flask(__name__)

@app.route("/")
def hello():
    # lb=load_workbook('jusik.xlsx')
    # ws=lb.active
    # lb.close()
    return render_template('hello.html')

@app.route("/save",methods=['POST','GET'])
def save():
    if request.method =='POST':
        if os.path.isfile('member.xlsx'):
            wb=load_workbook('member.xlsx')
        else:
            wb=Workbook()
        ws=wb.active
        ws.append([request.form['aa'],request.form['bb']])
        wb.save('member.xlsx')
        wb.close()
    name="홍길동"
    aaa=[10,20,30,40]
    return render_template('save.html',name=name,aaa=aaa) 

@app.route("/dusrma",methods=['POST','GET'])
def dusrma():
    insu=0
    if request.method=='POST':
        insu=request.form['won']
    result=ds.dusrma(insu)
    return render_template("dusrma.html",result=result)


if __name__=='__main__':
    app.run(debug=True)